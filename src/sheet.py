"""작업지를 읽는다. manifest.csv 를 엑셀에서 고쳐 쓰는 것이 기본 흐름이다.

organize가 전표에서 사실(거래일·금액·승인번호·가맹점)을 뽑아 채우고,
뒤쪽 네 칼럼(경비유형·비즈니스목적·코멘트·참석자)은 사람이 보고 고친다.
그 파일을 그대로 fix_expenses에 넘기면 적힌 대로 Concur에 넣는다.

.csv 와 .xlsx 를 받는다. xlsx는 openpyxl이 있을 때만 된다.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

REQUIRED = ["거래일", "승인번호"]

# 숙박비는 상세에 넣을 것이 많다. 입실·퇴실로 날짜 범위를 만들고, 숙박위치와
# Booking Channel은 화면 드롭다운에서 고른 값이어야 한다.
LODGING_COLUMNS = ["입실날짜", "퇴실날짜", "숙박위치", "Booking Channel"]
EDITABLE = ["경비유형", "비즈니스목적", "코멘트", "참석자", *LODGING_COLUMNS]

DATE_COLUMNS = ["입실날짜", "퇴실날짜"]

# 금액 칼럼 이름. '합계'로 쓰던 시절의 작업지도 그대로 열리게 둘 다 받는다.
AMOUNT_COLUMNS = ["금액", "합계"]


@dataclass
class SheetRow:
    """match()가 Slip처럼 다룰 수 있게 when/amount/merchant 이름을 맞춘다."""

    when: date
    amount: int
    merchant: str
    approval: str
    type_name: str
    purpose: str
    comment: str
    attendee: str
    checkin: date | None = None
    checkout: date | None = None
    location: str = ""
    channel: str = ""

    @property
    def nights(self) -> int:
        """숙박일수. 8/2 입실 8/8 퇴실이면 6박이다."""
        if not (self.checkin and self.checkout):
            return 0
        return (self.checkout - self.checkin).days


class SheetError(Exception):
    """작업지를 믿고 쓸 수 없을 때."""


def nightly_split(amount: int, nights: int) -> list[int]:
    """숙박비를 하루치로 나눈다. 소수점 없이, 합은 정확히 원래 금액.

    Concur의 일일 객실 요금은 소수점을 받지 않는다. 1,000,000원 3박이면
    333,333.33이 되는데 그냥 버림하면 999,999원이 되어 1원이 빈다. 나머지를
    앞 날짜부터 1원씩 얹어서 합을 맞춘다.
    """
    if nights <= 0:
        raise SheetError(f"숙박일수가 {nights}입니다. 입실·퇴실 날짜를 확인해 주세요.")
    base, rest = divmod(amount, nights)
    return [base + 1] * rest + [base] * (nights - rest)


def _rows_from_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _rows_from_xlsx(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SheetError(
            "xlsx를 읽으려면 openpyxl이 필요합니다: pip install openpyxl\n"
            "또는 엑셀에서 CSV로 저장해서 넘겨 주세요."
        ) from None
    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    return [
        {h: ("" if v is None else str(v).strip()) for h, v in zip(header, r)} for r in rows[1:]
    ]


# 이 유형을 고르면 그 칸들을 채워야 한다. 초록으로 물들여 알린다.
ATTENDEE_REQUIRED_TYPE = "내부 직원간 식음료"
LODGING_TYPE = "숙박비"

# 날짜는 8/2 로 보이게 한다. 값 자체는 진짜 날짜라 정렬도 계산도 된다.
DATE_FORMAT = "m/d"

DATE_PATTERNS = ["%Y-%m-%d", "%Y/%m/%d", "%m/%d", "%Y.%m.%d"]


def _as_date(value) -> date | None:
    """엑셀 셀이나 문자열에서 날짜를 뽑는다. 못 읽으면 None."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()[:19]
    for pattern in DATE_PATTERNS:
        try:
            when = datetime.strptime(text[:10] if len(pattern) > 5 else text, pattern)
        except ValueError:
            continue
        return when.date()
    try:  # '2026-08-02 00:00:00' 처럼 시각이 붙어 오는 경우
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None

# 숫자로 넣고 천단위 구분을 붙일 칼럼. 75400 보다 75,400 이 읽기 쉽고,
# 엑셀에서 합계도 바로 낼 수 있다. 저장되는 값은 그대로 숫자라 다시 읽는 데
# 영향이 없다 - 쉼표는 표시 형식일 뿐이다.
MONEY_FORMAT = "#,##0"

# 칸 너비. 엑셀의 너비 단위는 기본 글꼴의 '0' 한 글자 폭이라, 한글은 두 칸으로
# 세야 글자가 잘리지 않는다. 드롭다운 화살표와 여백으로 조금 더 준다.
WIDTH_MIN, WIDTH_MAX, WIDTH_PAD = 9, 42, 3


def _display_len(value: str) -> int:
    """한글·한자는 두 칸, 나머지는 한 칸으로 센다."""
    return sum(2 if ord(ch) > 0x2E7F else 1 for ch in value)


def _width(header: str, values: list) -> float:
    """머리글과 값 중 가장 긴 것에 맞춘다."""
    longest = max(
        [_display_len(header)] + [_display_len(str(v)) for v in values if v not in (None, "")]
    )
    return max(WIDTH_MIN, min(WIDTH_MAX, longest + WIDTH_PAD))


# 유형을 고르면 따라 채워질 값. 사람이 유형을 고르기 전에는 무엇이 맞는지 알 수
# 없으므로 값이 아니라 수식을 넣는다. 숙박비를 고르면 나타나고, 다른 유형이면
# 빈 칸으로 남는다. 사람이 드롭다운에서 다른 값을 고르면 수식은 그 값으로 덮인다.
TYPE_DEFAULTS = {LODGING_TYPE: {"숙박위치": "국내", "Booking Channel": "Others"}}


def write_xlsx(columns: list[str], rows: list[dict], path: Path,
               choices: dict[str, list[str]], hidden: list[str] | None = None) -> None:
    """드롭다운으로 고를 칸에 목록을 걸어서 내보낸다.

    choices는 {칼럼 이름: 고를 수 있는 값들}. 목록을 수식에 직접 넣으면 255자
    제한에 걸린다(한글 유형명이 길다). 목록마다 숨긴 시트를 만들어 참조한다.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise SheetError("xlsx를 만들려면 openpyxl이 필요합니다: pip install openpyxl") from None

    money = next((c for c in AMOUNT_COLUMNS if c in columns), None)
    money_at = columns.index(money) if money else -1
    date_at = {columns.index(c) for c in DATE_COLUMNS if c in columns}

    wb = Workbook()
    ws = wb.active
    ws.title = "전표"
    ws.append(columns)
    for r in rows:
        ws.append([r.get(c, "") for c in columns])
        line = ws.max_row
        if money_at >= 0:
            cell = ws.cell(row=line, column=money_at + 1)
            try:
                cell.value = int(str(cell.value).replace(",", "").strip())
            except (TypeError, ValueError):
                pass  # 숫자가 아니면 그대로 둔다. 사람이 보고 고칠 일이다
            else:
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right")
        # 날짜 칸은 비어 있어도 서식을 걸어둔다. 8/2 라고 치면 날짜가 되고
        # 화면에도 8/2 로 남는다. 연도는 엑셀이 올해로 잡는다.
        for at in date_at:
            cell = ws.cell(row=line, column=at + 1)
            cell.value = _as_date(cell.value)
            cell.number_format = DATE_FORMAT

    type_col = get_column_letter(columns.index("경비유형") + 1)
    for line in range(2, len(rows) + 2):
        for type_name, defaults in TYPE_DEFAULTS.items():
            for name, value in defaults.items():
                if name not in columns:
                    continue
                cell = ws.cell(row=line, column=columns.index(name) + 1)
                if cell.value in (None, ""):
                    cell.value = f'=IF(${type_col}{line}="{type_name}","{value}","")'

    for name, options in choices.items():
        if name not in columns or not options:
            continue
        ref = wb.create_sheet(f"목록_{name}"[:31])
        for option in options:
            ref.append([option])
        ref.sheet_state = "hidden"

        col = get_column_letter(columns.index(name) + 1)
        dv = DataValidation(
            type="list",
            formula1=f"='{ref.title}'!$A$1:$A${len(options)}",
            allow_blank=True,  # 빈 칸은 '이 값은 건드리지 마라'는 뜻이라 허용한다
            showDropDown=False,  # False가 화살표를 '보이게' 한다 (openpyxl의 뜻이 반대다)
            showErrorMessage=True,
            errorStyle="stop",  # 경고가 아니라 거부. 목록 밖의 값은 아예 못 넣는다
        )
        dv.error = "목록에서 골라 주세요. 직접 입력하실 수 없습니다."
        dv.errorTitle = name
        dv.prompt = "목록에서 골라 주세요. 비워두시면 이 값은 건드리지 않습니다."
        dv.promptTitle = name
        dv.showInputMessage = True
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{len(rows) + 1}")

    col = get_column_letter(columns.index("경비유형") + 1)

    # 그 유형에서 꼭 채워야 하는 칸을 초록으로 칠한다. 빈 칸은 눈에 안 띄어서
    # 빠뜨리기 쉽다. 식음료면 참석자, 숙박비면 입실·퇴실·숙박위치·채널이다.
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    needed = {ATTENDEE_REQUIRED_TYPE: ["참석자"], LODGING_TYPE: LODGING_COLUMNS}
    for type_name, names in needed.items():
        for name in names:
            if name not in columns or not rows:
                continue
            letter = get_column_letter(columns.index(name) + 1)
            ws.conditional_formatting.add(
                f"{letter}2:{letter}{len(rows) + 1}",
                FormulaRule(formula=[f'${col}2="{type_name}"'], fill=green),
            )

    for i, name in enumerate(columns, 1):
        dim = ws.column_dimensions[get_column_letter(i)]
        dim.width = _width(name, [r.get(name, "") for r in rows])
        if hidden and name in hidden:
            dim.hidden = True
    ws.freeze_panes = "A2"
    # 머리글에 필터를 걸어둔다. 날짜나 유형별로 골라 보기 좋다.
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
    wb.save(path)


def load(path: Path) -> list[SheetRow]:
    if not path.exists():
        raise SheetError(f"작업지가 없습니다: {path}\n먼저 B단계(파싱 · 작업지 생성)를 실행해 주세요.")
    raw = _rows_from_xlsx(path) if path.suffix.lower() == ".xlsx" else _rows_from_csv(path)
    if not raw:
        raise SheetError(f"작업지가 비어 있습니다: {path}")

    missing = [c for c in REQUIRED if c not in raw[0]]
    money = next((c for c in AMOUNT_COLUMNS if c in raw[0]), None)
    if money is None:
        missing.append(AMOUNT_COLUMNS[0])
    if missing:
        raise SheetError(f"작업지에 다음 칸이 없습니다: {', '.join(missing)} ({path})")

    out = []
    for i, r in enumerate(raw, 2):  # 2행부터 (1행은 머리글)
        if not (r.get("승인번호") or "").strip():
            continue
        try:
            when = datetime.strptime(str(r["거래일"]).strip()[:10], "%Y-%m-%d").date()
            amount = int(float(str(r[money]).replace(",", "").strip()))
        except ValueError as exc:
            raise SheetError(f"{path} {i}행의 거래일/{money}을 읽지 못했습니다: {exc}") from None
        checkin, checkout = _as_date(r.get("입실날짜")), _as_date(r.get("퇴실날짜"))
        if bool(checkin) != bool(checkout):
            raise SheetError(f"{path} {i}행: 입실날짜와 퇴실날짜는 둘 다 적어 주세요.")
        if checkin and checkout <= checkin:
            raise SheetError(
                f"{path} {i}행: 퇴실날짜({checkout})가 입실날짜({checkin})보다 뒤여야 합니다."
            )
        out.append(
            SheetRow(
                when=when,
                amount=amount,
                merchant=(r.get("가맹점명") or "").strip(),
                approval=r["승인번호"].strip(),
                type_name=(r.get("경비유형") or "").strip(),
                purpose=(r.get("비즈니스목적") or "").strip(),
                comment=(r.get("코멘트") or "").strip(),
                attendee=(r.get("참석자") or "").strip(),
                checkin=checkin,
                checkout=checkout,
                location=(r.get("숙박위치") or "").strip(),
                channel=(r.get("Booking Channel") or "").strip(),
            )
        )
    return out
