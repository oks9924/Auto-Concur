"""숙박비 계산 회귀 테스트.

일일 객실 요금의 합이 경비 금액과 1원이라도 다르면 Concur가 저장을 막거나,
저장돼도 나중에 정산에서 걸린다. 그 규칙을 여기서 고정한다.
"""

from datetime import date

import pytest

from src import sheet
from src.fix_expenses import Lodging
from src.sheet import nightly_split


@pytest.mark.parametrize(
    "amount,nights",
    [(1_000_000, 3), (999_999, 6), (75_400, 1), (100, 7), (350_000, 2), (1, 1)],
)
def test_합은_언제나_금액과_같다(amount, nights):
    per = nightly_split(amount, nights)
    assert len(per) == nights
    assert sum(per) == amount


def test_나머지는_앞_날짜부터_1원씩():
    # 1,000,000 / 3 = 333,333.33. 버리면 1원이 빈다.
    assert nightly_split(1_000_000, 3) == [333_334, 333_333, 333_333]


def test_소수점은_들어가지_않는다():
    assert all(isinstance(x, int) for x in nightly_split(1_000_000, 7))


def test_숙박일수가_0이면_멈춘다():
    with pytest.raises(sheet.SheetError, match="숙박일수"):
        nightly_split(100_000, 0)


def test_입실_퇴실로_숙박일수를_센다():
    # 8/2 입실, 8/8 퇴실이면 6박이다.
    stay = Lodging(date(2026, 8, 2), date(2026, 8, 8), "울산", "직접 예약")
    assert stay.nights == 6
    assert stay.dates() == [date(2026, 8, d) for d in range(2, 8)]


def test_명세_날짜와_금액_개수가_맞는다():
    stay = Lodging(date(2026, 8, 2), date(2026, 8, 8), "", "")
    assert len(stay.dates()) == len(nightly_split(1_000_000, stay.nights))


def test_퇴실이_입실보다_앞이면_작업지에서_멈춘다(tmp_path):
    path = tmp_path / "m.csv"
    path.write_text(
        "거래일,금액,승인번호,입실날짜,퇴실날짜\n2026-08-02,900000,1,2026-08-08,2026-08-02\n",
        encoding="utf-8-sig",
    )
    with pytest.raises(sheet.SheetError, match="퇴실날짜"):
        sheet.load(path)


def test_한쪽만_적으면_멈춘다(tmp_path):
    path = tmp_path / "m.csv"
    path.write_text(
        "거래일,금액,승인번호,입실날짜,퇴실날짜\n2026-08-02,900000,1,2026-08-02,\n",
        encoding="utf-8-sig",
    )
    with pytest.raises(sheet.SheetError, match="둘 다"):
        sheet.load(path)


def test_엑셀에서_8슬래시2로_적어도_읽는다(tmp_path):
    from src.organize import MANIFEST_COLUMNS

    rows = [
        {
            "거래일": "2026-08-02", "금액": "900000", "승인번호": "1",
            "경비유형": "숙박비", "입실날짜": "2026-08-02", "퇴실날짜": "2026-08-08",
            "숙박위치": "울산", "Booking Channel": "직접 예약",
        }
    ]
    path = tmp_path / "m.xlsx"
    sheet.write_xlsx(MANIFEST_COLUMNS, rows, path, {"경비유형": ["숙박비"]})

    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    ws = load_workbook(path)["전표"]
    cell = ws.cell(row=2, column=MANIFEST_COLUMNS.index("입실날짜") + 1)
    assert cell.value.date() == date(2026, 8, 2)
    assert cell.number_format == sheet.DATE_FORMAT  # 화면에는 8/2 로 보인다
    assert get_column_letter(MANIFEST_COLUMNS.index("퇴실날짜") + 1)

    loaded = sheet.load(path)[0]
    assert loaded.nights == 6
    assert loaded.location == "울산" and loaded.channel == "직접 예약"


def test_빈_항목별_명세는_추가_버튼을_누른다():
    """'항목별 명세 없음' 화면에는 '반복' 콤보박스가 없다. 버튼부터 눌러야 생긴다."""
    from src.fix_expenses import itemization_step

    화면 = {"form": False, "add": True, "empty": True}
    assert itemization_step(화면) == "add"


def test_이미_있는_명세는_추가를_누르지_않는다():
    """명세가 있을 때도 '추가' 버튼은 있다. 누르면 필요 없는 명세가 하나 더 생긴다."""
    from src.fix_expenses import itemization_step

    화면 = {"form": False, "add": True, "empty": False}
    assert itemization_step(화면) == "skip"


def test_입력_폼이_떠_있으면_바로_채운다():
    from src.fix_expenses import itemization_step

    assert itemization_step({"form": True, "add": True, "empty": False}) == "fill"


def test_1박은_반복_칸을_찾지_않는다():
    """반복할 밤이 없어서 '일일 금액 동일/다름' 칸이 아예 안 나온다."""
    from src.fix_expenses import needs_recurrence

    assert not needs_recurrence(1, lambda: False)
    assert nightly_split(120_000, 1) == [120_000]  # 한 행이면 동일이든 다름이든 같다


def test_1박이라도_칸이_있으면_고른다():
    from src.fix_expenses import needs_recurrence

    assert needs_recurrence(1, lambda: True)


def test_여러_박이면_화면을_보지_않고_고른다():
    """아직 안 그려졌을 때 미리 보고 '없다'고 판단하면 안 된다."""
    from src.fix_expenses import needs_recurrence

    def 보면안됨():
        raise AssertionError("여러 박이면 화면을 볼 것도 없이 골라야 한다")

    assert needs_recurrence(6, 보면안됨)


def test_화면_상태는_보이는_글자로_읽는다():
    """셀렉터를 추측하지 않는다. 화면에 뜨는 글자가 근거다."""
    from src.fix_expenses import (
        ADD_ITEMIZATION_JS,
        ADD_ITEMIZATION_TEXT,
        EMPTY_ITEMIZATION_TEXT,
        ITEMIZATION_STATE_JS,
    )

    assert EMPTY_ITEMIZATION_TEXT in ITEMIZATION_STATE_JS
    assert ADD_ITEMIZATION_TEXT in ITEMIZATION_STATE_JS
    assert ADD_ITEMIZATION_TEXT in ADD_ITEMIZATION_JS


def test_숨은_저장_버튼은_뒤로_밀되_버리지_않는다():
    """실측 두 번.

    1) 화면 밖의 exp-save-expense-hidden 이 먼저 걸려 30초를 기다리다 실패했다.
    2) 그래서 목록에서 아예 뺐더니 '저장 버튼을 찾지 못했습니다'로 그냥
       넘어가버렸다. 빼지 말고 순서만 뒤로 미는 것이 맞다.
    """
    from src.fix_expenses import SAVE_BUTTONS_JS

    assert "save-hidden-button" in SAVE_BUTTONS_JS
    assert "found.sort" in SAVE_BUTTONS_JS  # 거르는 게 아니라 순서만 민다
    assert "return found.map" in SAVE_BUTTONS_JS  # 후보를 전부 돌려준다


def test_숙박비인데_날짜가_없으면_짚어준다():
    # 코멘트만 넣고 지나가면 다 된 것처럼 보인다. 빠진 값을 알려줘야 한다.
    from src.fix_expenses import _gaps

    class Entry:
        checkin = checkout = None
        location = channel = attendee = ""

    holes = _gaps(Entry(), "숙박비")
    assert "입실·퇴실 날짜" in holes and "숙박 위치" in holes


def test_식음료인데_참석자가_없으면_짚어준다():
    from src.fix_expenses import _gaps

    class Entry:
        checkin = checkout = None
        location = channel = attendee = ""

    assert _gaps(Entry(), "내부 직원간 식음료") == ["참석자"]


def test_숙박비를_고른_행에만_기본값이_나온다(tmp_path):
    """유형을 고르기 전에는 무엇이 맞는지 모른다. 그래서 값이 아니라 수식을 넣는다."""
    from openpyxl import load_workbook

    from src import settings
    from src.organize import MANIFEST_COLUMNS
    from src.settings import type_defaults

    base = dict.fromkeys(MANIFEST_COLUMNS, "")
    rows = [base | {"거래일": "2026-08-02", "금액": "450000", "승인번호": "1"}]
    path = tmp_path / "m.xlsx"
    sheet.write_xlsx(MANIFEST_COLUMNS, rows, path, settings.choices(settings.DEFAULTS),
                     type_defaults=settings.type_defaults(settings.DEFAULTS))

    ws = load_workbook(path)["전표"]
    셀 = ws.cell(row=2, column=MANIFEST_COLUMNS.index("숙박위치") + 1)
    assert 셀.value == '=IF($M2="숙박비","국내","")'
    셀 = ws.cell(row=2, column=MANIFEST_COLUMNS.index("Booking Channel") + 1)
    assert 셀.value == '=IF($M2="숙박비","Others","")'

    # 기본값은 드롭다운 목록 안에 있어야 한다. 아니면 사람이 고칠 수도 없다.
    고를수있는값 = settings.choices(settings.DEFAULTS)
    기본값 = type_defaults(settings.DEFAULTS)["숙박비"]
    assert 기본값["숙박위치"] in 고를수있는값["숙박위치"]
    assert 기본값["Booking Channel"] in 고를수있는값["Booking Channel"]


def test_사람이_적은_값은_수식으로_덮이지_않는다(tmp_path):
    from openpyxl import load_workbook

    from src import settings
    from src.organize import MANIFEST_COLUMNS

    base = dict.fromkeys(MANIFEST_COLUMNS, "")
    rows = [base | {"거래일": "2026-08-02", "금액": "450000", "승인번호": "1", "숙박위치": "해외"}]
    path = tmp_path / "m.xlsx"
    sheet.write_xlsx(MANIFEST_COLUMNS, rows, path, settings.choices(settings.DEFAULTS),
                     type_defaults=settings.type_defaults(settings.DEFAULTS))

    ws = load_workbook(path)["전표"]
    assert ws.cell(row=2, column=MANIFEST_COLUMNS.index("숙박위치") + 1).value == "해외"


def test_칸이_비어_있으면_설정_기본값을_쓴다(tmp_path):
    """엑셀 수식이 계산되지 않은 채 저장되면 빈 칸으로 읽힌다. 그때도 채워야 한다."""
    import csv

    from src import fix_expenses as fx
    from src import settings
    from src.attach_receipts import Row

    cols = ["거래일", "금액", "승인번호", "경비유형", "입실날짜", "퇴실날짜", "숙박위치",
            "Booking Channel", "비즈니스목적", "코멘트", "참석자"]
    path = tmp_path / "m.csv"
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        w.writerow({"거래일": "2026-08-02", "금액": "450000", "승인번호": "1",
                    "경비유형": "숙박비", "입실날짜": "2026-08-02", "퇴실날짜": "2026-08-08",
                    "숙박위치": "", "Booking Channel": "", "비즈니스목적": "",
                    "코멘트": "", "참석자": ""})

    screen = [Row(0, date(2026, 8, 2), 450000, "숙박비", "ID1", "숙박비", "HOTEL")]
    plans, _, _ = fx.plans_from_sheet(settings.DEFAULTS, screen, path, 1)
    stay = plans[0][0].lodging
    assert stay.location == "국내" and stay.channel == "Others"
