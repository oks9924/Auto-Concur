"""이미 처리된 건이 섞여 있을 때의 규칙.

두 번째 실행이 첫 번째와 다르게 굴어야 한다. 영수증은 두 번 붙이면 안 되고,
유형·목적·코멘트는 작업지 값으로 다시 맞춰야 하고, 참석자는 이미 맞으면
건드리지 않아야 한다.
"""

from datetime import date

from src.attach_receipts import Row, Slip, match
from src.fix_expenses import name_matches


def row(index: int, amount: int, receipt=None) -> Row:
    return Row(index, date(2026, 7, 1), amount, "", f"ID{index}", "", "", receipt)


def slip(amount: int, approval: str) -> Slip:
    from pathlib import Path

    return Slip(Path("a.pdf"), date(2026, 7, 1), amount, "가맹점", approval)


def test_영수증이_있는_행을_구분한다():
    # 첨부 여부는 attach_phase가 이 값으로 가른다.
    rows = [row(0, 1000, True), row(1, 2000, False), row(2, 3000, None)]
    assert [r.has_receipt for r in rows] == [True, False, None]


def test_영수증_유무는_매칭에_영향을_주지_않는다():
    # 붙었는지는 매칭 다음에 가른다. 매칭에서 빼면 짝이 밀린다.
    pairs, skipped = match([slip(1000, "1")], [row(0, 1000, True)], 1)
    assert len(pairs) == 1 and not skipped


def test_참석자_이름_대조():
    # 작업지는 'kyungsik.oh', 화면은 'Oh Kyungsik' 이다.
    assert name_matches("kyungsik.oh", "Oh Kyungsik")
    assert name_matches("kyungsik.oh", "Oh Kyungsik (kyungsik.oh@x.com)")
    assert not name_matches("kyungsik.oh", "Kim Minsu")
    assert not name_matches("", "Oh Kyungsik")  # 빈 검색어는 아무나 맞추면 안 된다


def test_화면_이름과_작업지_검색어를_맞춘다():
    # 실측: 화면 이름은 'Oh Kyungsik', 'Lee Kyungmin'. 작업지는 'kyungsik.oh'.
    from src.fix_expenses import name_matches

    화면 = ["Lee Kyungmin", "Oh Kyungsik"]
    작업지 = ["kyungsik.oh"]

    지울사람 = [n for n in 화면 if not any(name_matches(q, n) for q in 작업지)]
    넣을사람 = [q for q in 작업지 if not any(name_matches(q, n) for n in 화면)]
    assert 지울사람 == ["Lee Kyungmin"]
    assert 넣을사람 == []


def test_이름이_비슷해도_다른_사람은_안_맞는다():
    from src.fix_expenses import name_matches

    # kyungsik 과 kyungmin 은 다른 사람이다
    assert not name_matches("kyungsik.oh", "Lee Kyungmin")
    assert not name_matches("kyungmin.lee", "Oh Kyungsik")


def _한건(tmp_path, 참석자, 추가, 유형="내부 직원간 식음료"):
    import csv

    from src import fix_expenses as fx
    from src import settings
    from src.attach_receipts import Row

    cols = ["거래일", "금액", "승인번호", "경비유형", "참석자", "추가 참석자",
            "비즈니스목적", "코멘트"]
    path = tmp_path / "m.csv"
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        w.writerow({"거래일": "2026-07-02", "금액": "13500", "승인번호": "1",
                    "경비유형": 유형, "참석자": 참석자, "추가 참석자": 추가,
                    "비즈니스목적": "", "코멘트": "점심"})
    cfg = {**settings.DEFAULTS, "attendee_default": "kyungsik.oh"}
    screen = [Row(0, date(2026, 7, 2), 13500, "", "ID1", "환급 불가", "CAFE")]
    plans, gaps, _ = fx.plans_from_sheet(cfg, screen, path, 1)
    return plans[0][0].attendee, gaps


def test_참석자와_추가_참석자를_합친다(tmp_path):
    누구, _ = _한건(tmp_path, "kyungsik.oh", "hong.gildong")
    assert 누구 == "kyungsik.oh, hong.gildong"


def test_참석자를_지우면_추가_참석자만_넣는다(tmp_path):
    """수식을 지운 것은 본인을 빼겠다는 뜻이다."""
    누구, _ = _한건(tmp_path, "", "hong.gildong, kim.minsu")
    assert 누구 == "hong.gildong, kim.minsu"


def test_둘_다_비면_설정값을_쓴다(tmp_path):
    """아무도 안 넣으면 식음료는 필수값이 비어서 리포트가 안 나간다."""
    누구, gaps = _한건(tmp_path, "", "")
    assert 누구 == "kyungsik.oh"
    assert not gaps


def test_같은_이름이_양쪽에_있으면_한_번만(tmp_path):
    누구, _ = _한건(tmp_path, "kyungsik.oh", "kyungsik.oh, kim.minsu")
    assert 누구 == "kyungsik.oh, kim.minsu"


def test_식음료가_아니면_넣지_않는다(tmp_path):
    """참석자 칸은 그 유형에만 있다. 다른 유형에 넣으려 하면 버튼이 없어 실패한다."""
    누구, _ = _한건(tmp_path, "kyungsik.oh", "hong.gildong", "주차비")
    assert 누구 == ""


def test_참석자는_수식_추가_참석자는_빈_칸(tmp_path):
    """참석자는 자동으로 채워지고, 사람이 손대는 칸은 '추가 참석자' 다."""
    from openpyxl import load_workbook

    from src import settings, sheet
    from src.organize import MANIFEST_COLUMNS

    cfg = {**settings.DEFAULTS, "attendee_default": "kyungsik.oh"}
    base = dict.fromkeys(MANIFEST_COLUMNS, "")
    rows = [base | {"거래일": "2026-07-02", "금액": "13500", "승인번호": "1"}]
    path = tmp_path / "m.xlsx"
    sheet.write_xlsx(MANIFEST_COLUMNS, rows, path, settings.choices(cfg),
                     type_defaults=settings.type_defaults(cfg))

    ws = load_workbook(path)["전표"]
    참석자 = ws.cell(row=2, column=MANIFEST_COLUMNS.index("참석자") + 1)
    추가 = ws.cell(row=2, column=MANIFEST_COLUMNS.index("추가 참석자") + 1)
    assert 참석자.value == '=IF($M2="내부 직원간 식음료","kyungsik.oh","")'
    assert 추가.value in (None, "")


def test_추가_참석자_칸에_설명이_붙는다(tmp_path):
    """마우스를 올리면 뜨는 메모와, 칸을 고르면 뜨는 쪽지 둘 다 있어야 한다."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    from src import settings, sheet
    from src.organize import MANIFEST_COLUMNS

    base = dict.fromkeys(MANIFEST_COLUMNS, "")
    rows = [base | {"거래일": "2026-07-02", "금액": "13500", "승인번호": "1"}]
    path = tmp_path / "m.xlsx"
    sheet.write_xlsx(MANIFEST_COLUMNS, rows, path, settings.choices(settings.DEFAULTS))

    ws = load_workbook(path)["전표"]
    col = get_column_letter(MANIFEST_COLUMNS.index("추가 참석자") + 1)
    assert "빈칸으로 두세요" in ws[f"{col}1"].comment.text

    쪽지 = [d for d in ws.data_validations.dataValidation
            if d.promptTitle == "추가 참석자"]
    assert 쪽지 and "빈칸으로 두세요" in 쪽지[0].prompt


def test_초록은_참석자와_추가_참석자_둘_다(tmp_path):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    from src import settings, sheet
    from src.organize import MANIFEST_COLUMNS

    base = dict.fromkeys(MANIFEST_COLUMNS, "")
    rows = [base | {"거래일": "2026-07-02", "금액": "13500", "승인번호": "1"}]
    path = tmp_path / "m.xlsx"
    sheet.write_xlsx(MANIFEST_COLUMNS, rows, path, settings.choices(settings.DEFAULTS))

    ws = load_workbook(path)["전표"]
    칠한칸 = {str(r.sqref) for r in ws.conditional_formatting}
    for name in ("참석자", "추가 참석자"):
        assert f"{get_column_letter(MANIFEST_COLUMNS.index(name) + 1)}2" in 칠한칸


def test_검색_결과_고르는_규칙이_파이썬과_같다():
    """화면에서 고를 때 쓰는 JS와 지울지 판단할 때 쓰는 파이썬이 어긋나면 안 된다.

    JS 쪽은 브라우저에서만 돌아서 여기서는 규칙만 견준다.
    """
    from src.fix_expenses import MATCH_NAME_FN, name_matches

    # 같은 규칙이어야 한다: 검색어를 토막 내서 전부 들어 있으면 같은 사람
    assert "every(p => low.includes(p))" in MATCH_NAME_FN
    assert name_matches("kyungsik.oh", "Oh Kyungsik")
    assert not name_matches("kyungsik.oh", "Lee Kyungmin")


def test_검색_결과를_검색어로_고른다():
    """첫 번째를 고르면 결과가 여럿일 때 엉뚱한 사람이 들어간다."""
    from src.fix_expenses import SELECT_ATTENDEE_OPTION_JS

    assert SELECT_ATTENDEE_OPTION_JS.startswith("(q) =>")
    assert "findOption(q)" in SELECT_ATTENDEE_OPTION_JS


def test_건너뛴_근거가_어디에_있는지_말한다(tmp_path, capsys):
    """'이미 붙인 2건' 은 Concur를 본 것이 아니라 우리가 남긴 기록이다.

    화면에 영수증이 없어도 그 파일에 적혀 있으면 건너뛴다. 어디를 봐야
    하는지 말해주지 않으면 '왜 하나만 붙었냐'가 된다.
    """
    import csv

    from src import attach_receipts as ar
    from src.organize import MANIFEST_COLUMNS

    (tmp_path / "a.pdf").write_bytes(b"%PDF-1.4\n")
    with (tmp_path / "manifest.csv").open("w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=MANIFEST_COLUMNS)
        w.writeheader()
        w.writerow({**dict.fromkeys(MANIFEST_COLUMNS, ""), "파일명": "a.pdf",
                    "거래일": "2026-08-09", "금액": "17000", "승인번호": "111",
                    "가맹점명": "가게"})
    ar.done_path(tmp_path).write_text("111\n", encoding="utf-8")

    assert ar.attach_phase(None, "http://x", tmp_path, True, 1, None) == 0
    말 = capsys.readouterr().out
    assert ar.done_path(tmp_path).name in 말  # 어디에 적혀 있는지
    assert "--again" in 말  # 어떻게 다시 붙이는지


def test_값이_안_채워진_목록은_다_읽힌_것이_아니다():
    """행 껍데기는 먼저 그려지고 날짜·금액은 조금 뒤에 채워진다.

    그 사이에 읽으면 행 수는 맞는데 값이 비어서, 작업지의 모든 줄이
    '후보 없음'이 된다 - 짝은 날짜와 금액으로만 짓기 때문이다.
    """
    from src.fix_expenses import rows_ready

    빈행 = [Row(0, None, None, "", "ID1", "", ""), Row(1, None, None, "", "ID2", "", "")]
    assert not rows_ready(빈행)
    assert not rows_ready([])

    반쯤 = [*빈행, Row(2, date(2026, 8, 9), 17000, "", "ID3", "", "")]
    assert rows_ready(반쯤)  # 하나라도 채워졌으면 그리는 중이 아니다


def test_값이_없는_행은_짝짓기에서_빠진다():
    """빠진 것을 말없이 빼면 '왜 못 찾았는지' 를 알 수 없다."""
    from src.attach_receipts import match

    화면 = [Row(0, None, None, "", "ID1", "", ""),
           Row(1, date(2026, 8, 9), 17000, "", "ID2", "", "")]
    쓸수있는것 = [r for r in 화면 if r.expense_id and r.when and r.amount]
    assert len(쓸수있는것) == 1

    pairs, missing = match([slip(19000, "1")], 쓸수있는것, 1)
    assert not pairs and missing[0][1] == "후보 없음"


def test_숨겨진_칸도_글자를_읽는다():
    """innerText 는 화면에 그려진 글자다. 칸이 숨겨져 있으면 빈 문자열이 나온다.

    실측 2026-08-11: 같은 리포트를 다시 열었더니 3건 다 날짜·금액이 비었고,
    10초를 기다려도 그대로였다. 그리는 중이 아니라 안 보이는 상태였다.
    """
    from src.attach_receipts import READ_ROWS_JS

    assert "el.innerText || el.textContent" in READ_ROWS_JS


def test_못_읽으면_행_마크업을_남긴다():
    """추측으로 셀렉터를 고치지 않는다. 실물을 보고 고친다."""
    import inspect

    from src import attach_receipts as ar
    from src import fix_expenses as fx

    assert "concur-rows.json" in inspect.getsource(ar.dump_rows)
    for func in (ar.attach_phase, fx.fix_phase):
        source = inspect.getsource(func)
        assert "읽지 못해" in source and "dump_rows(page)" in source


def test_못_읽은_칸의_글자를_보여준다(capsys):
    """'읽지 못했습니다'만 있으면 비었는지 형식이 다른지 알 수 없다.

    화면 말이 영어면 날짜가 '08/09/2026' 으로 나오는데 우리는 '2026-08-09'
    만 읽는다. 그 차이는 이 줄이 있어야 드러난다.
    """
    from src.attach_receipts import Row, print_unreadable

    화면 = [
        Row(0, None, None, "", "ID1", "", "", None, "08/09/2026", "KRW 17,000.00"),
        Row(1, date(2026, 8, 9), 19000, "", "ID2", "", "", None, "2026-08-09", "19,000"),
    ]
    print_unreadable(화면)
    말 = capsys.readouterr().out
    assert "08/09/2026" in 말 and "KRW 17,000.00" in 말
    assert "2026-08-09" not in 말  # 읽힌 행은 굳이 늘어놓지 않는다


def test_다_읽혔으면_아무_말도_하지_않는다(capsys):
    from src.attach_receipts import Row, print_unreadable

    print_unreadable([Row(0, date(2026, 8, 9), 17000, "", "ID1", "", "")])
    assert capsys.readouterr().out == ""


def test_행_요약에서_금액을_읽는다():
    """실측 2026-08-11: 금액 칸(amount-cell) 훅이 안 잡혀 3건 다 0원이었다.

    행마다 화면낭독기용 요약이 하나 붙어 있고 거기에 금액이 들어 있다.
    칸 이름이 바뀌어도 이건 남는다. 실물(concur-rows) 그대로 가져왔다.
    """
    from src.attach_receipts import amount_from_label

    # KRW 뒤는 &nbsp;(\xa0) 다. 보통 공백이 아니다.
    요약 = "경비, 내부 직원간 식음료 (점심, 야근식대, 부서회식, 음료 등), KRW\xa017,000, 날짜, 2026-08-09 선택"
    assert amount_from_label(요약) == 17000
    assert amount_from_label("경비, 숙박비, KRW 764,707, 날짜, 2026-07-16 선택") == 764707
    assert amount_from_label("경비, 주차비, 날짜, 2026-07-16 선택") is None
    assert amount_from_label("") is None


def test_금액_칸이_읽히면_그걸_쓴다():
    """요약은 대비책이다. 칸이 읽히는 동안은 칸이 우선이다."""
    from src.attach_receipts import READ_ROWS_JS

    assert "pick(r, 'amount-cell')" in READ_ROWS_JS
    assert "screen-reader-only" in READ_ROWS_JS


def test_영수증_판정은_미리보기_버튼만_본다():
    """실측 2026-08-11 concur-receipts.json. 세 행의 영수증 칸이 이랬다:

      row0  data-nuiexp="receipt-thumbnail-button-20260809_17000_00817287.pdf"  -> 붙음
      row1  data-nuiexp="rcpt-btn-attach-receipt" (올리기 아이콘)                -> 안 붙음
      row2  data-nuiexp="receipt-thumbnail-button-20260716_764707_00815079.pdf" -> 붙음

    예전 규칙은 '칸 안에 img/svg/button/a 가 있으면 붙음'이었다. row1의 올리기
    버튼도 button이라 붙음으로 셌고, 그래서 19,000원 영수증이 영영 안 붙었다.
    """
    from src.attach_receipts import READ_ROWS_JS

    assert '[data-nuiexp^="receipt-thumbnail-button"]' in READ_ROWS_JS
    # 헐거운 옛 규칙이 남아 있으면 안 된다
    assert "querySelector('img, svg, button, a')" not in READ_ROWS_JS


def test_건너뛴_이유에_영수증_파일_이름을_적는다():
    """붙어 있는 파일 이름이 작업지와 다르면 엉뚱한 영수증이 붙어 있는 것이다."""
    import inspect

    from src import attach_receipts as ar

    소스 = inspect.getsource(ar.attach_phase)
    assert "row.receipt_file" in 소스
